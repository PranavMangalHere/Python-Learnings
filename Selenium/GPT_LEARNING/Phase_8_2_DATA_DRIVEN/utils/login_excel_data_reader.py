from openpyxl import load_workbook
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(BASE_DIR, "test_data/", "login_test_data.xlsx")

def read_excel(file_path, sheet_name):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    data = []
    headers = [cell.values for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)
    return data


def get_login_data():
    data = read_excel(file_path, "LoginData")

    return [
        (row["username"], row["password"], row["expected"])
        for row in data
    ]

def get_test_ids():
    data = read_excel(file_path, "LoginData")

    return [row["test_name"] for row in data]
